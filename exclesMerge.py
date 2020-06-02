# -*- coding:utf-8 -*-
'''
@author = 随时静听 - xuqiu
@date = 2020-06-01 22:57
@product_name = PyCharm
@project_name = KeySearch
@filename = exclesMerge.py
'''

import os
from argparse import ArgumentParser
import openpyxl
import json
import glob
import string

## 读取 excel的每一行，忽略ignore个行，ignore设置表头
def readXlsx(filename,ignore=1):
    if not os.path.exists(filename):
        yield None
    try:
        wb= openpyxl.load_workbook(filename)
    except:
        print ("[-] XLSX load Failed" + filename)
        yield  None
    sheet=wb.get_sheet_by_name(wb.sheetnames[0])
    for i,row in enumerate(sheet.rows):
        if i<1:
            continue
        yield row

##获取文件列表
def getAllfile(path,ext=".xlsx"):
    if os.path.exists(path) and os.path.isdir(path):
        fileLst=glob.glob1(path,"*"+ext)
        fileLst= map( lambda _:os.path.join(path,_),fileLst)
        return fileLst
    else:
        print("[Error] Get xlsx file failed! Path is not exists!"+path)
        return []


# 读取配置文件写入头并设置格式,添加title
def addTitle(ws,conf="./data/title.conf"):
    uppercase=string.uppercase
    style={"bold":True,"italic":True}
    if os.path.exists(conf):
        with open(conf,"r") as f:
            titles=f.read()
            titles=titles.split("\n")
            titles=filter( lambda x:x.strip(),titles)
            titles=map(lambda x:x.split("::"),titles)
            for i,line in enumerate(titles):
                ws.cell(row=1,column=i+1,value=line[1])
                #字体样式
                fontstryle=openpyxl.styles.Font(name=u"宋体",size=int(line[2]),bold=style.get(line[3].lower(),False), italic=style.get(line[4].lower(),False),color=line[4]\
                                         )
                ws[uppercase[i]+"1"].font = fontstryle
                ws[uppercase[i] + "1"]=line[1]
                ws[uppercase[i] + "1"].fill = openpyxl.styles.PatternFill(fill_type='solid',fgColor=line[5])
                #对齐方式
                ws[uppercase[i]+"1"].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                #设置行高和列宽
                ws.row_dimensions[1].height=int(line[6])
                ws.column_dimensions[uppercase[i]].width=int(line[7])
    return  ws

def makeMegre(files,output,ignore=1,titles="./data/title.conf"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws = addTitle(ws,titles)
    r = 2
    border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin", color="000000"), \
                                    right=openpyxl.styles.Side(border_style="thin", color="000000"), \
                                    top=openpyxl.styles.Side(border_style="thin", color="000000"), \
                                  bottom=openpyxl.styles.Side(border_style="thin", color="000000"))
    failedlst=[]
    with open("failed.log","a+") as f:

        for filename in files:
            print u"[Info] Handing file ",filename,"..."
            try:
                for row in readXlsx(filename,ignore):
                    for ci,c in enumerate(row):
                        if ci==0:
                            ws.cell(row=r, column= 1, value=r-1).border=border
                            continue
                        ws.cell(row=r,column=ci+1,value=c.value).border=border
                    r=r+1
            except Exception,e:
                f.write(filename+"\n")
                failedlst.append(filename)
                print u"[Error] 文件读取失败 ",filename ,e.message
    wb.save(output)
    return failedlst


def main():
    parser = ArgumentParser(
        description=
        '''
        Excels Merge tool!\n
        '''
    )
    parser.add_argument(
        '-n',
        '--num',
        dest='num',
        type=int,
        nargs='?',
        default=1,
        help=u'忽略Excel文件表头的行数，默认值为1')
    parser.add_argument(
        '-o',
        '--output',
        dest='output',
        nargs='?',
        default='./reports/outut.xlsx',
        help=u'指定导出文件的路径，默认值为 ./reports/output.xlsx'
    )
    parser.add_argument(
        '-i',
        '--input',
        dest='input',
        nargs='?',
        required=True,
        help=u'指定需要处理的多个Excel文件路径'
    )
    args = parser.parse_args()

    filelst = getAllfile(args.input)
    print "[Info] Find xlsx file num:" + str(len(filelst))

    output=args.output
    if os.path.exists(output):
        print u"[Error] 文件已存在：",output
        print u"[Info] 程序已退出！"
        return
    else:
        filepath=os.path.dirname(output)
        if not os.path.exists(filepath):
            os.makedirs(filepath)
        output=os.path.abspath(output)
    ignoreNum=args.num
    failedlst=makeMegre(filelst,output,ignoreNum)
    total = len(filelst)
    failednum = len(failedlst)
    print u"[Info] 检测到处理文件:", str(total), u"处理成功:", str(total - failednum), u"处理失败：", str(failednum)
    print u"[Info] 导出路径为：", output

if __name__ == "__main__":
    main()
    pass